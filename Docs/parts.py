import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Define the file path
file_path = r'C:\Users\stefa\Downloads\Parts.txt'

# Read the file
with open(file_path, 'r') as file:
    lines = file.readlines()

# Skip the header and parse the data
data = []
for line in lines[1:]:  # Skip the header line
    if line.strip():  # Avoid empty lines
        name, qty = line.split('\t')
        name = name.strip()
        
        # Apply naming convention and strip prefixes
        if name.startswith('RFM') or name == 'DW01A':
            part_type = 'Odd Part'  # RFM and DW01A go to Odd Parts
            footprint = ''  # These parts may not have a common footprint
        elif name.startswith('R'):
            if name.startswith('RV'):
                part_type = 'Variable Resistor'
                name = name[2:]  # Remove the 'RV' prefix
                footprint = ''  # Variable resistors may not have a common footprint
            else:
                part_type = 'Resistor'
                name = name[1:]  # Remove the 'R' prefix
                footprint = '0805'
        elif name.startswith('C'):
            part_type = 'Capacitor'
            name = name[1:]  # Remove the 'C' prefix
            footprint = '0805'
        elif name.startswith('IND'):
            part_type = 'Inductor'
            name = name[3:]  # Remove the 'IND' prefix
            footprint = ''  # Inductors may not have a common footprint
        elif name.startswith('D'):
            part_type = 'Diode'
            name = name[1:]  # Remove the 'D' prefix
            footprint = ''  # Diodes may not have a common footprint
        elif name.lower() == 'led' or name.lower() == 'led':
            part_type = 'LED'
            name = 'LED'
            footprint = '0805'
        else:
            part_type = 'Odd Part'  # Everything else is classified as an "Odd part"
            footprint = ''  # Odd parts may not have a common footprint
        
        data.append([part_type, name.strip(), int(qty.strip()), footprint])

# Create a DataFrame
df = pd.DataFrame(data, columns=["Category", "Name", "QTY", "Footprint"])

# Group by Category and Name, then sum the quantities
df = df.groupby(["Category", "Name", "Footprint"], as_index=False).sum()

# Define the custom sort order
sort_order = {
    'Odd Part': 1,
    'Resistor': 2,
    'Variable Resistor': 3,
    'Capacitor': 4,
    'Inductor': 5,
    'Diode': 6,
    'LED': 7
}

# Apply custom sort order to the DataFrame
df['SortOrder'] = df['Category'].map(sort_order)
df = df.sort_values(by=['SortOrder', 'Name', 'QTY'])

# Drop the SortOrder column as it's no longer needed
df = df.drop(columns=['SortOrder'])

# Define the output Excel file path (same as the input path with .xlsx extension)
output_file = file_path.replace('.txt', '.xlsx')

# Save the DataFrame to an Excel file
df.to_excel(output_file, index=False)

# Load the workbook and sheet for formatting
wb = load_workbook(output_file)
ws = wb.active

# Adjust column widths and merge category cells
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Define color fills for each category
category_colors = {
    'Odd Part': 'FFFF99',  # Light Yellow
    'Resistor': 'FFCCCC',  # Light Red
    'Variable Resistor': 'FF9966',  # Light Orange
    'Capacitor': '99CCFF',  # Light Blue
    'Inductor': 'CCFFCC',  # Light Green
    'Diode': 'FFCC99',  # Light Peach
    'LED': 'FF99CC'  # Light Pink
}

# Merge cells for each category and apply color
current_category = None
start_row = 2  # Skip header row

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1), start=2):
    cell = row[0]
    if cell.value != current_category:
        if current_category is not None:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=i-1, end_column=1)
            ws.cell(row=start_row, column=1).alignment = Alignment(vertical="center", horizontal="center")
            # Apply color
            for r in range(start_row, i):
                for col in range(1, 5):
                    ws.cell(row=r, column=col).fill = PatternFill(start_color=category_colors[current_category], end_color=category_colors[current_category], fill_type="solid")
        current_category = cell.value
        start_row = i
    elif i == ws.max_row:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=i, end_column=1)
        ws.cell(row=start_row, column=1).alignment = Alignment(vertical="center", horizontal="center")
        # Apply color
        for r in range(start_row, i + 1):
            for col in range(1, 5):
                ws.cell(row=r, column=col).fill = PatternFill(start_color=category_colors[current_category], end_color=category_colors[current_category], fill_type="solid")

# Save the formatted Excel file
wb.save(output_file)

print(f"Excel file saved as: {output_file}")
